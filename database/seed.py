import os
from datetime import datetime, timedelta

from openpyxl import load_workbook

from database.db import get_db_connection

def import_excel_tickets_to_db():
    file_path = os.path.join(os.path.dirname(__file__), "Help-Desk-Excel.xlsx")

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    sheet = workbook["Ticket_Log"]

    connection = get_db_connection()
    cursor = connection.cursor()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        ticket_id, date_opened, date_closed, department, category, priority, status, technician, resolution_time, sla_met, root_cause, notes = row

        if not ticket_id:
            continue

        cursor.execute("""
            INSERT INTO tickets
            (name, department, issue, priority, category, status, submitted_at, closed_at, assigned_to, completed_by, notes, sla_due_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            "Generated User",
            department,
            notes if notes else "Generated ticket",
            priority,
            category,
            status,
            str(date_opened),
            str(date_closed) if date_closed else "",
            technician if technician else "Unassigned",
            technician if status == "Closed" else "",
            root_cause if root_cause else "",
            ""
        ))

    connection.commit()
    connection.close()
    workbook.close()

    print("Excel tickets imported into SQLite successfully.")

def add_ticket_to_excel(name, department, issue, priority, category, status, submitted_at, assigned_to):
    try:
        file_path = os.path.join(os.path.dirname(__file__), "Help-Desk-Excel.xlsx")

        workbook = load_workbook(file_path)
        sheet = workbook["Ticket_Log"]

        next_row = sheet.max_row + 1

        sheet.cell(row=next_row, column=1).value = f"T-{next_row - 1:03d}"
        sheet.cell(row=next_row, column=2).value = submitted_at
        sheet.cell(row=next_row, column=3).value = ""
        sheet.cell(row=next_row, column=4).value = department
        sheet.cell(row=next_row, column=5).value = category
        sheet.cell(row=next_row, column=6).value = priority
        sheet.cell(row=next_row, column=7).value = status
        sheet.cell(row=next_row, column=8).value = assigned_to
        sheet.cell(row=next_row, column=11).value = ""
        sheet.cell(row=next_row, column=12).value = issue

        workbook.save(file_path)
        workbook.close()

        print("Ticket saved to Excel successfully.")

    except Exception as error:
        print("Excel save failed:", error)
        
    
def backfill_sla_data():
    connection = get_db_connection()
    cursor = connection.cursor()

    cursor.execute("SELECT * FROM tickets WHERE status = 'Closed'")
    tickets = cursor.fetchall()

    for ticket in tickets:
        if ticket["resolution_time"] and ticket["sla_met"]:
            continue

        try:
            submitted_time = datetime.strptime(ticket["submitted_at"], "%Y-%m-%d %H:%M:%S")
        except Exception:
            try:
                submitted_time = datetime.strptime(ticket["submitted_at"], "%m/%d/%Y %I:%M %p")
            except Exception:
                continue

        try:
            closed_time = datetime.strptime(ticket["closed_at"], "%Y-%m-%d %H:%M:%S")
        except Exception:
            try:
                closed_time = datetime.strptime(ticket["closed_at"], "%m/%d/%Y %I:%M %p")
            except Exception:
                continue

        if ticket["priority"] == "High":
            sla_due = submitted_time + timedelta(hours=1)
        elif ticket["priority"] == "Medium":
            sla_due = submitted_time + timedelta(hours=8)
        else:
            sla_due = submitted_time + timedelta(hours=24)

        difference = closed_time - submitted_time
        total_minutes = int(difference.total_seconds() / 60)
        hours = total_minutes // 60
        minutes = total_minutes % 60

        resolution_time = f"{hours} hr {minutes} min"
        sla_met = "Yes" if closed_time <= sla_due else "No"
        sla_due_at = sla_due.strftime("%m/%d/%Y %I:%M %p")

        cursor.execute("""
            UPDATE tickets
            SET resolution_time = ?,
                sla_met = ?,
                sla_due_at = ?
            WHERE id = ?
        """, (resolution_time, sla_met, sla_due_at, ticket["id"]))

    connection.commit()
    connection.close()

    print("Old SLA data updated.")