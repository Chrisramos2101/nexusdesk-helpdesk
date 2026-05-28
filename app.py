from flask import Flask, render_template, request, redirect, session
from datetime import datetime, timedelta
import sqlite3
import os
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev_secret_key")

tickets = []

def init_db():
    connection = sqlite3.connect("helpdesk.db")
    cursor = connection.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS tickets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            department TEXT NOT NULL,
            issue TEXT NOT NULL,
            priority TEXT NOT NULL,
            status TEXT NOT NULL,
            submitted_at TEXT NOT NULL,
            closed_at TEXT,
            assigned_to TEXT,
            completed_by TEXT,
            notes TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL
        )
    """)

    try:
        cursor.execute("ALTER TABLE tickets ADD COLUMN assigned_to TEXT DEFAULT 'Unassigned'")
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE tickets ADD COLUMN completed_by TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE tickets ADD COLUMN category TEXT DEFAULT 'Other'")
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE tickets ADD COLUMN sla_due_at TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE tickets ADD COLUMN submitted_by TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE tickets ADD COLUMN resolution_time TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE tickets ADD COLUMN sla_met TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass

    cursor.execute("SELECT * FROM users WHERE username = ?", ("admin",))
    admin_exists = cursor.fetchone()

    if not admin_exists:
        cursor.execute("""
            INSERT INTO users (username, password, role)
            VALUES (?, ?, ?)
        """, ("admin", "admin123", "admin"))

    cursor.execute("SELECT * FROM users WHERE username = ?", ("employee",))
    employee_exists = cursor.fetchone()

    if not employee_exists:
        cursor.execute("""
            INSERT INTO users (username, password, role)
            VALUES (?, ?, ?)
        """, ("employee", "employee123", "employee"))

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ticket_notes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticket_id INTEGER NOT NULL,
            note TEXT NOT NULL,
            created_by TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(ticket_id) REFERENCES tickets(id)
        )
    """)

    connection.commit()
    connection.close()

def import_excel_tickets_to_db():
    file_path = os.path.join(os.path.dirname(__file__), "Help-Desk-Excel.xlsx")

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    sheet = workbook["Ticket_Log"]

    connection = sqlite3.connect("helpdesk.db")
    cursor = connection.cursor()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        ticket_id, date_opened, date_closed, department, category, priority, status, technician, resolution_time, sla_met, root_cause, notes = row

        if not ticket_id:
            continue

        cursor.execute("""
            INSERT INTO tickets
            (name, department, issue, priority, category, status, submitted_at, closed_at, assigned_to, completed_by, notes, sla_due_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            "Generated User",
            department,
            notes if notes else "Generated ticket",
            priority,
            category,
            status,
            str(date_opened),
            str(date_closed) if date_closed else "",
            technician if technician else "Unassigned",
            technician if status == "Closed" else "",
            root_cause if root_cause else "",
            ""
        ))

    connection.commit()
    connection.close()
    workbook.close()

    print("Excel tickets imported into SQLite successfully.")

def add_ticket_to_excel(name, department, issue, priority, category, status, submitted_at, assigned_to):
    try:
        file_path = os.path.join(os.path.dirname(__file__), "Help-Desk-Excel.xlsx")

        workbook = load_workbook(file_path)
        sheet = workbook["Ticket_Log"]

        next_row = sheet.max_row + 1

        sheet.cell(row=next_row, column=1).value = f"T-{next_row - 1:03d}"
        sheet.cell(row=next_row, column=2).value = submitted_at
        sheet.cell(row=next_row, column=3).value = ""
        sheet.cell(row=next_row, column=4).value = department
        sheet.cell(row=next_row, column=5).value = category
        sheet.cell(row=next_row, column=6).value = priority
        sheet.cell(row=next_row, column=7).value = status
        sheet.cell(row=next_row, column=8).value = assigned_to
        sheet.cell(row=next_row, column=11).value = ""
        sheet.cell(row=next_row, column=12).value = issue

        workbook.save(file_path)
        workbook.close()

        print("Ticket saved to Excel successfully.")

    except Exception as error:
        print("Excel save failed:", error)
        

@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""

    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        connection = sqlite3.connect("helpdesk.db")
        connection.row_factory = sqlite3.Row
        cursor = connection.cursor()

        cursor.execute("""
            SELECT *
            FROM users
            WHERE username = ? AND password = ?
        """, (username, password))

        user = cursor.fetchone()
        connection.close()

        if user:
            session["username"] = user["username"]
            session["role"] = user["role"]
            return redirect("/")

        error = "Invalid username or password"

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


@app.route("/")
def home():
    if "username" not in session:
        return redirect("/login")

    connection = sqlite3.connect("helpdesk.db")
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()

    cursor.execute("SELECT COUNT(*) AS total FROM tickets WHERE status = 'Open'")
    open_tickets = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM tickets WHERE priority = 'High'")
    high_priority_tickets = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT sla_due_at
        FROM tickets
        WHERE status != 'Closed' AND sla_due_at != ''
    """)
    rows = cursor.fetchall()

    overdue_tickets = 0

    for row in rows:
        try:
            due_time = datetime.strptime(row["sla_due_at"], "%m/%d/%Y %I:%M %p")
            if datetime.now() > due_time:
                overdue_tickets += 1
        except Exception:
            pass

    connection.close()

    submitted = request.args.get("submitted")
    ticket_id = request.args.get("ticket_id")

    return render_template(
        "index.html",
        role=session["role"],
        open_tickets=open_tickets,
        overdue_tickets=overdue_tickets,
        high_priority_tickets=high_priority_tickets,
        submitted=submitted,
        ticket_id=ticket_id
    )


@app.route("/my_tickets")
def my_tickets():
    if "username" not in session:
        return redirect("/login")

    connection = sqlite3.connect("helpdesk.db")
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()

    cursor.execute("""
        SELECT *
        FROM tickets
        WHERE submitted_by = ?
        ORDER BY id DESC
    """, (session["username"],))

    tickets = cursor.fetchall()
    connection.close()

    return render_template("my_tickets.html", tickets=tickets)


@app.route("/submit", methods=["POST"])
def submit():
    name = request.form["name"]
    department = request.form["department"]
    issue = request.form["issue"]
    priority = request.form["priority"]
    category = request.form["category"]
    submitted_at = datetime.now().strftime("%m/%d/%Y %I:%M %p")
    submitted_by = session["username"]

    now = datetime.now()

    if priority == "High":
        sla_due_at = (now + timedelta(hours=1)).strftime("%m/%d/%Y %I:%M %p")
    elif priority == "Medium":
        sla_due_at = (now + timedelta(hours=8)).strftime("%m/%d/%Y %I:%M %p")
    else:
        sla_due_at = (now + timedelta(hours=24)).strftime("%m/%d/%Y %I:%M %p")

    connection = sqlite3.connect("helpdesk.db")
    cursor = connection.cursor()

    cursor.execute("""
        INSERT INTO tickets
        (name, department, issue, priority, category, status, submitted_at, closed_at, assigned_to, completed_by, notes, sla_due_at, submitted_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (name, department, issue, priority, category, "Open", submitted_at, "", "Unassigned", "", "", sla_due_at, submitted_by))

    connection.commit()
    new_ticket_id = cursor.lastrowid
    connection.close()

    return redirect(f"/?submitted=success&ticket_id={new_ticket_id}")

    add_ticket_to_excel(
        name,
        department,
        issue,
        priority,
        category,
        "Open",
        submitted_at,
        "Unassigned"
    )

    return redirect("/dashboard")

def get_sla_status(ticket):
    if ticket["status"] == "Closed":
        return "Completed"

    if not ticket["sla_due_at"]:
        return "No SLA"

    due_time = datetime.strptime(ticket["sla_due_at"], "%m/%d/%Y %I:%M %p")
    now = datetime.now()

    if now > due_time:
        return "Overdue"

    minutes_left = (due_time - now).total_seconds() / 60

    if minutes_left <= 60:
        return "Due Soon"

    return "On Track"

def calculate_resolution_time(submitted_at, closed_at):
    try:
        submitted_time = datetime.strptime(submitted_at, "%m/%d/%Y %I:%M %p")
        closed_time = datetime.strptime(closed_at, "%m/%d/%Y %I:%M %p")

        difference = closed_time - submitted_time
        total_minutes = int(difference.total_seconds() / 60)

        hours = total_minutes // 60
        minutes = total_minutes % 60

        return f"{hours} hr {minutes} min"

    except Exception:
        return ""


def calculate_sla_met(closed_at, sla_due_at):
    try:
        if not closed_at or not sla_due_at:
            return ""

        closed_time = datetime.strptime(closed_at, "%m/%d/%Y %I:%M %p")
        due_time = datetime.strptime(sla_due_at, "%m/%d/%Y %I:%M %p")

        if closed_time <= due_time:
            return "Yes"
        else:
            return "No"

    except Exception:
        return ""
    
def backfill_sla_data():
    connection = sqlite3.connect("helpdesk.db")
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()

    cursor.execute("SELECT * FROM tickets WHERE status = 'Closed'")
    tickets = cursor.fetchall()

    for ticket in tickets:
        if ticket["resolution_time"] and ticket["sla_met"]:
            continue

        try:
            submitted_time = datetime.strptime(ticket["submitted_at"], "%Y-%m-%d %H:%M:%S")
        except Exception:
            try:
                submitted_time = datetime.strptime(ticket["submitted_at"], "%m/%d/%Y %I:%M %p")
            except Exception:
                continue

        try:
            closed_time = datetime.strptime(ticket["closed_at"], "%Y-%m-%d %H:%M:%S")
        except Exception:
            try:
                closed_time = datetime.strptime(ticket["closed_at"], "%m/%d/%Y %I:%M %p")
            except Exception:
                continue

        if ticket["priority"] == "High":
            sla_due = submitted_time + timedelta(hours=1)
        elif ticket["priority"] == "Medium":
            sla_due = submitted_time + timedelta(hours=8)
        else:
            sla_due = submitted_time + timedelta(hours=24)

        difference = closed_time - submitted_time
        total_minutes = int(difference.total_seconds() / 60)
        hours = total_minutes // 60
        minutes = total_minutes % 60

        resolution_time = f"{hours} hr {minutes} min"
        sla_met = "Yes" if closed_time <= sla_due else "No"
        sla_due_at = sla_due.strftime("%m/%d/%Y %I:%M %p")

        cursor.execute("""
            UPDATE tickets
            SET resolution_time = ?,
                sla_met = ?,
                sla_due_at = ?
            WHERE id = ?
        """, (resolution_time, sla_met, sla_due_at, ticket["id"]))

    connection.commit()
    connection.close()

    print("Old SLA data updated.")

@app.route("/dashboard")
def dashboard():
    if "username" not in session:
        return redirect("/login")
    
    if session["role"] != "admin":
        return redirect("/")
    
    selected_status = request.args.get("status", "Open")
    search_query = request.args.get("search", "")
    selected_priority = request.args.get("priority", "")
    selected_category = request.args.get("category", "")
    selected_assigned = request.args.get("assigned_to", "")

    priority_order = {
        "High": 1,
        "Medium": 2,
        "Low": 3
    }

    connection = sqlite3.connect("helpdesk.db")
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()

    query = """
        SELECT * FROM tickets
        WHERE status = ?
    """

    params = [selected_status]

    if search_query:
        query += """
            AND (
                name LIKE ?
                OR department LIKE ?
                OR issue LIKE ?
            )
        """
        search_pattern = f"%{search_query}%"
        params.extend([search_pattern, search_pattern, search_pattern])

    if selected_priority:
        query += " AND priority = ?"
        params.append(selected_priority)

    if selected_category:
        query += " AND category = ?"
        params.append(selected_category)

    if selected_assigned:
        query += " AND assigned_to LIKE ?"
        params.append(f"%{selected_assigned}%")

    cursor.execute(query, params)

    tickets_from_db = cursor.fetchall()

    cursor.execute("""
        SELECT * FROM ticket_notes
        ORDER BY created_at DESC
    """)

    notes_from_db = cursor.fetchall()

    tickets_with_sla = []

    for ticket in tickets_from_db:
        ticket_dict = dict(ticket)
        ticket_dict["sla_status"] = get_sla_status(ticket)
        tickets_with_sla.append(ticket_dict)

    notes_by_ticket = {}

    for note in notes_from_db:
        ticket_id = note["ticket_id"]

        if ticket_id not in notes_by_ticket:
            notes_by_ticket[ticket_id] = []

        notes_by_ticket[ticket_id].append(note)

    cursor.execute("SELECT COUNT(*) AS total FROM tickets WHERE status = 'Open'")
    open_tickets = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM tickets WHERE status = 'In Progress'")
    in_progress_tickets = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM tickets WHERE status = 'Closed'")
    closed_tickets = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM tickets WHERE sla_met = 'Yes'")
    sla_met_count = cursor.fetchone()["total"]

    overdue_tickets = 0

    cursor.execute("""
        SELECT sla_due_at
        FROM tickets
        WHERE status != 'Closed' AND sla_due_at != ''
    """)

    open_sla_rows = cursor.fetchall()

    for row in open_sla_rows:
        try:
            due_time = datetime.strptime(row["sla_due_at"], "%m/%d/%Y %I:%M %p")
            if datetime.now() > due_time:
                overdue_tickets += 1
        except Exception:
            pass

    cursor.execute("""
        SELECT assigned_to, COUNT(*) AS total
        FROM tickets
        WHERE status != 'Closed'
        GROUP BY assigned_to
    """)
    tech_load = cursor.fetchall()

    connection.close()

    sla_order = {
        "Overdue": 1,
        "Due Soon": 2,
        "On Track": 3,
        "No SLA": 4,
        "Completed": 5
    }

    sorted_tickets = sorted(
        tickets_with_sla,
        key=lambda ticket: (
            sla_order[ticket["sla_status"]],
            priority_order[ticket["priority"]]
        )
    )

    return render_template(
        "dashboard.html",
        tickets=sorted_tickets,
        selected_status=selected_status,
        notes_by_ticket=notes_by_ticket,
        admin_name=session["username"],
        search_query=search_query,
        selected_priority=selected_priority,
        selected_category=selected_category,
        selected_assigned=selected_assigned,
        open_tickets=open_tickets,
        in_progress_tickets=in_progress_tickets,
        closed_tickets=closed_tickets,
        overdue_tickets=overdue_tickets,
        sla_met_count=sla_met_count,
        tech_load=tech_load,
    )

@app.route("/update_ticket/<int:ticket_id>", methods=["POST"])
def update_ticket(ticket_id):
    if "username" not in session or session["role"] != "admin":
        return redirect("/login")

    name = request.form["name"]
    department = request.form["department"]
    issue = request.form["issue"]
    priority = request.form["priority"]
    category = request.form["category"]
    status = request.form["status"]
    assigned_to = request.form["assigned_to"]
    new_note = request.form["notes"]

    connection = sqlite3.connect("helpdesk.db")
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()

    cursor.execute("SELECT * FROM tickets WHERE id = ?", (ticket_id,))
    ticket = cursor.fetchone()

    closed_at = ticket["closed_at"] if ticket["closed_at"] else ""
    completed_by = ticket["completed_by"] if ticket["completed_by"] else ""
    resolution_time = ticket["resolution_time"] if "resolution_time" in ticket.keys() else ""
    sla_met = ticket["sla_met"] if "sla_met" in ticket.keys() else ""

    if status == "Closed" and not closed_at:
        closed_at = datetime.now().strftime("%m/%d/%Y %I:%M %p")
        completed_by = assigned_to if assigned_to != "Unassigned" else session["username"]
        resolution_time = calculate_resolution_time(ticket["submitted_at"], closed_at)
        sla_met = calculate_sla_met(closed_at, ticket["sla_due_at"])

    if status != "Closed":
        closed_at = ""
        completed_by = ""
        resolution_time = ""
        sla_met = ""

    cursor.execute("""
        UPDATE tickets
        SET name = ?,
            department = ?,
            issue = ?,
            priority = ?,
            category = ?,
            status = ?,
            assigned_to = ?,
            closed_at = ?,
            completed_by = ?,
            resolution_time = ?,
            sla_met = ?
        WHERE id = ?
    """, (
        name,
        department,
        issue,
        priority,
        category,
        status,
        assigned_to,
        closed_at,
        completed_by,
        resolution_time,
        sla_met,
        ticket_id
    ))

    if new_note.strip() != "":
        note_time = datetime.now().strftime("%m/%d/%Y %I:%M %p")

        cursor.execute("""
            INSERT INTO ticket_notes (ticket_id, note, created_by, created_at)
            VALUES (?, ?, ?, ?)
        """, (ticket_id, new_note, session["username"], note_time))

    connection.commit()
    connection.close()

    return redirect(f"/dashboard?status={status}")


@app.route("/assign_me/<int:ticket_id>", methods=["POST"])
def assign_me(ticket_id):
    if "username" not in session or session["role"] != "admin":
        return redirect("/login")

    connection = sqlite3.connect("helpdesk.db")
    cursor = connection.cursor()

    cursor.execute("""
        UPDATE tickets
        SET assigned_to = ?
        WHERE id = ?
    """, (session["username"].capitalize(), ticket_id))

    connection.commit()
    connection.close()

    return redirect("/dashboard")


@app.route("/quick_close/<int:ticket_id>", methods=["POST"])
def quick_close(ticket_id):
    if "username" not in session or session["role"] != "admin":
        return redirect("/login")

    closed_at = datetime.now().strftime("%m/%d/%Y %I:%M %p")

    connection = sqlite3.connect("helpdesk.db")
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()

    cursor.execute("SELECT * FROM tickets WHERE id = ?", (ticket_id,))
    ticket = cursor.fetchone()

    resolution_time = calculate_resolution_time(ticket["submitted_at"], closed_at)
    sla_met = calculate_sla_met(closed_at, ticket["sla_due_at"])

    cursor.execute("""
        UPDATE tickets
        SET status = 'Closed',
            closed_at = ?,
            completed_by = ?,
            resolution_time = ?,
            sla_met = ?
        WHERE id = ?
    """, (
        closed_at,
        ticket["assigned_to"] if ticket["assigned_to"] else session["username"],
        resolution_time,
        sla_met,
        ticket_id
    ))

    connection.commit()
    connection.close()

    return redirect("/dashboard?status=Closed")


@app.route("/delete_ticket/<int:ticket_id>", methods=["POST"])
def delete_ticket(ticket_id):
    if "username" not in session or session["role"] != "admin":
        return redirect("/login")

    connection = sqlite3.connect("helpdesk.db")
    cursor = connection.cursor()

    cursor.execute("DELETE FROM tickets WHERE id = ?", (ticket_id,))

    connection.commit()
    connection.close()

    return redirect("/dashboard")

@app.route("/stats")
def stats():
    if "username" not in session:
        return redirect("/login")

    if session["role"] != "admin":
        return redirect("/")

    connection = sqlite3.connect("helpdesk.db")
    connection.row_factory = sqlite3.Row
    cursor = connection.cursor()

    cursor.execute("SELECT COUNT(*) AS total FROM tickets")
    total_tickets = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM tickets WHERE status = 'Open'")
    open_tickets = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM tickets WHERE status = 'In Progress'")
    in_progress_tickets = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM tickets WHERE status = 'Closed'")
    closed_tickets = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM tickets WHERE sla_met = 'Yes'")
    sla_met_count = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM tickets WHERE sla_met = 'No'")
    sla_missed_count = cursor.fetchone()["total"]

    cursor.execute("""
        SELECT assigned_to, COUNT(*) AS total
        FROM tickets
        WHERE sla_met = 'No'
        GROUP BY assigned_to
    """)
    sla_missed_by_tech = cursor.fetchall()

    cursor.execute("""
        SELECT category, COUNT(*) AS total
        FROM tickets
        GROUP BY category
    """)
    category_stats = cursor.fetchall()

    cursor.execute("SELECT COUNT(*) AS total FROM tickets WHERE priority = 'High'")
    high_priority_tickets = cursor.fetchone()["total"]

    cursor.execute("SELECT department, COUNT(*) AS total FROM tickets GROUP BY department")
    department_stats = cursor.fetchall()

    cursor.execute("SELECT priority, COUNT(*) AS total FROM tickets GROUP BY priority")
    priority_stats = cursor.fetchall()

    cursor.execute("SELECT category, COUNT(*) AS total FROM tickets GROUP BY category")
    category_stats = cursor.fetchall()

    cursor.execute("""
        SELECT assigned_to, COUNT(*) AS total
        FROM tickets
        WHERE assigned_to IS NOT NULL AND assigned_to != ''
        GROUP BY assigned_to
    """)
    assigned_stats = cursor.fetchall()

    cursor.execute("""
        SELECT completed_by, COUNT(*) AS total
        FROM tickets
        WHERE status = 'Closed' AND completed_by != ''
        GROUP BY completed_by
    """)
    completed_by_stats = cursor.fetchall()

    cursor.execute("""
        SELECT submitted_at, COUNT(*) AS total
        FROM tickets
        GROUP BY submitted_at
        ORDER BY submitted_at DESC
        LIMIT 7
    """)
    daily_stats = cursor.fetchall()

    open_percent = round((open_tickets / total_tickets) * 100, 1) if total_tickets else 0
    progress_percent = round((in_progress_tickets / total_tickets) * 100, 1) if total_tickets else 0
    closed_percent = round((closed_tickets / total_tickets) * 100, 1) if total_tickets else 0

    priority_stats = [
        {
            "priority": row["priority"],
            "total": row["total"],
            "percent": round((row["total"] / total_tickets) * 100, 1) if total_tickets else 0
        }
        for row in priority_stats
    ]

    department_stats = [
        {
            "department": row["department"],
            "total": row["total"],
            "percent": round((row["total"] / total_tickets) * 100, 1) if total_tickets else 0
        }
        for row in department_stats
    ]

    category_stats = [
        {
            "category": row["category"],
            "total": row["total"],
            "percent": round((row["total"] / total_tickets) * 100, 1) if total_tickets else 0
        }
        for row in category_stats
    ]

    assigned_stats = [
        {
            "assigned_to": row["assigned_to"],
            "total": row["total"],
            "percent": round((row["total"] / total_tickets) * 100, 1) if total_tickets else 0
        }
        for row in assigned_stats
    ]

    cursor.execute("""
    SELECT COUNT(*) AS total
    FROM tickets
    WHERE status != 'Closed'
    AND sla_due_at != ''
    """)
    potential_sla_tickets = cursor.fetchone()["total"]

    overdue_tickets = 0

    cursor.execute("""
        SELECT sla_due_at
        FROM tickets
        WHERE status != 'Closed'
        AND sla_due_at != ''
    """)

    open_sla_rows = cursor.fetchall()

    for row in open_sla_rows:
        try:
            due_time = datetime.strptime(row["sla_due_at"], "%m/%d/%Y %I:%M %p")
            if datetime.now() > due_time:
                overdue_tickets += 1
        except Exception:
            pass

    connection.close()

    return render_template(
        "stats.html",
        total_tickets=total_tickets,
        open_tickets=open_tickets,
        in_progress_tickets=in_progress_tickets,
        closed_tickets=closed_tickets,
        high_priority_tickets=high_priority_tickets,
        open_percent=open_percent,
        progress_percent=progress_percent,
        closed_percent=closed_percent,
        priority_stats=priority_stats,
        department_stats=department_stats,
        category_stats=category_stats,
        assigned_stats=assigned_stats,
        completed_by_stats=completed_by_stats,
        daily_stats=daily_stats,
        sla_met_count=sla_met_count,
        sla_missed_count=sla_missed_count,
        sla_missed_by_tech=sla_missed_by_tech,
        overdue_tickets=overdue_tickets,
    )

@app.route("/healthz")
def healthz():
    return "OK", 200

if __name__ == "__main__":
    init_db()
    app.run()